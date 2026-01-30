#!/usr/bin/env python3
"""
PDF Batch Processor
Processes multiple PDF files in a folder to:
1. Count three-letter codes from the last line of each page across all PDFs
2. Combine all PDFs (excluding those with "multi-page" in filename) into one,
   sorted alphabetically by the first line of each page
"""

import pdfplumber
from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook
import sys
import os
import re
from collections import Counter
from pathlib import Path

def extract_three_letter_code(line):
    """
    Extract the three-letter code from a line.
    Expected format: "xxx MM/DD/YY cccc"
    """
    pattern = r'^([a-zA-Z0-9]{3})\s+\d{1,2}/\d{1,2}/\d{2,4}\s+[a-zA-Z]{4}\s*$'
    match = re.match(pattern, line.strip())
    if match:
        return match.group(1)
    return None

def get_first_line(page):
    """
    Extract the first non-empty line from a page.
    
    Args:
        page: pdfplumber page object
        
    Returns:
        First non-empty line as string, or empty string if none found
    """
    text = page.extract_text()
    if text:
        lines = text.split('\n')
        for line in lines:
            if line.strip():
                return line.strip()
    return ""

def count_codes_in_folder(folder_path, verbose=False):
    """
    Count three-letter codes from all PDFs in a folder.
    
    Args:
        folder_path: Path to the folder containing PDFs
        verbose: If True, print details about each file
        
    Returns:
        Counter object with code counts
    """
    code_counter = Counter()
    pdf_files = list(Path(folder_path).glob('*.pdf'))
    
    if not pdf_files:
        print(f"No PDF files found in {folder_path}")
        return code_counter
    
    print(f"Found {len(pdf_files)} PDF file(s)")
    print("=" * 80)
    
    for pdf_file in sorted(pdf_files):
        print(f"\nProcessing: {pdf_file.name}")
        
        try:
            with pdfplumber.open(pdf_file) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    text = page.extract_text()
                    
                    if text:
                        lines = text.split('\n')
                        
                        # Get last non-empty line
                        last_line = None
                        for line in reversed(lines):
                            if line.strip():
                                last_line = line.strip()
                                break
                        
                        if last_line:
                            code = extract_three_letter_code(last_line)
                            
                            if code:
                                code_counter[code] += 1
                                if verbose:
                                    print(f"  Page {page_num}: Found code '{code}'")
                            elif verbose:
                                print(f"  Page {page_num}: No code found in: {last_line}")
        
        except Exception as e:
            print(f"  Error processing {pdf_file.name}: {e}")
    
    print("=" * 80)
    return code_counter

def combine_pdfs_alphabetically(folder_path, output_path, verbose=False):
    """
    Combine PDFs from a folder, sorted alphabetically by first line of each page.
    Excludes files with "multi-page" in the filename.
    
    Args:
        folder_path: Path to the folder containing PDFs
        output_path: Path for the output combined PDF
        verbose: If True, print details about each page
        
    Returns:
        Number of pages in combined PDF
    """
    pdf_files = list(Path(folder_path).glob('*.pdf'))
    
    # Filter out files with "multi-page" in the name
    pdf_files = [f for f in pdf_files if "multi-page" not in f.name.lower()]
    
    if not pdf_files:
        print("No PDF files to combine (after filtering)")
        return 0
    
    print(f"\nCombining {len(pdf_files)} PDF file(s) (excluding 'multi-page' files)")
    print("=" * 80)
    
    # Store pages with their first line for sorting
    pages_with_keys = []
    
    for pdf_file in sorted(pdf_files):
        if verbose:
            print(f"Reading: {pdf_file.name}")
        
        try:
            # Use pdfplumber to extract text for sorting
            with pdfplumber.open(pdf_file) as plumber_pdf:
                # Use pypdf to get the actual pages for combining
                pypdf_reader = PdfReader(pdf_file)
                
                for page_num, plumber_page in enumerate(plumber_pdf.pages):
                    first_line = get_first_line(plumber_page)
                    pypdf_page = pypdf_reader.pages[page_num]
                    
                    pages_with_keys.append({
                        'first_line': first_line,
                        'page': pypdf_page,
                        'source': pdf_file.name,
                        'page_num': page_num + 1
                    })
                    
                    if verbose:
                        print(f"  Page {page_num + 1}: '{first_line[:50]}...'")
        
        except Exception as e:
            print(f"Error reading {pdf_file.name}: {e}")
    
    # Sort pages alphabetically by first line (case-insensitive)
    pages_with_keys.sort(key=lambda x: x['first_line'].lower())
    
    print(f"\nSorted {len(pages_with_keys)} pages alphabetically by first line")
    
    # Create the combined PDF
    writer = PdfWriter()
    
    for item in pages_with_keys:
        writer.add_page(item['page'])
    
    # Write the output
    with open(output_path, 'wb') as output_file:
        writer.write(output_file)
    
    print(f"Combined PDF saved to: {output_path}")
    print("=" * 80)
    
    return len(pages_with_keys)

def update_excel_spreadsheet(excel_path, code_counter, verbose=False):
    """
    Update Excel spreadsheet with code counts.
    Updates the second sheet, with codes in column A and counts in column D.
    
    Args:
        excel_path: Path to the Excel file
        code_counter: Counter object with code counts
        verbose: If True, print detailed information
        
    Returns:
        Number of codes updated
    """
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return 0
    
    print(f"\nUpdating Excel spreadsheet: {excel_path}")
    print("=" * 80)
    
    try:
        # Load the workbook
        wb = load_workbook(excel_path)
        
        # Get the second sheet (index 1)
        if len(wb.sheetnames) < 2:
            print("Error: Spreadsheet must have at least 2 sheets")
            return 0
        
        sheet = wb[wb.sheetnames[1]]
        sheet_name = wb.sheetnames[1]
        print(f"Working with sheet: '{sheet_name}'")
        
        # Build a dictionary of existing codes and their row positions
        existing_codes = {}
        row = 2  # Start from row 2 (assuming row 1 is header)
        
        # Scan existing codes in column A
        while True:
            cell_value = sheet[f'A{row}'].value
            if cell_value is None or cell_value == '':
                break
            existing_codes[str(cell_value).strip()] = row
            row += 1
        
        next_empty_row = row
        
        if verbose:
            print(f"Found {len(existing_codes)} existing codes in spreadsheet")
            print(f"Next empty row: {next_empty_row}")
        
        # Update or add each code
        updated_count = 0
        added_count = 0
        
        for code, count in sorted(code_counter.items()):
            if code in existing_codes:
                # Update existing code
                row_num = existing_codes[code]
                old_value = sheet[f'D{row_num}'].value or 0
                sheet[f'D{row_num}'] = count
                updated_count += 1
                
                if verbose:
                    print(f"  Updated '{code}' at row {row_num}: {old_value} → {count}")
            else:
                # Add new code
                sheet[f'A{next_empty_row}'] = code
                sheet[f'D{next_empty_row}'] = count
                added_count += 1
                
                if verbose:
                    print(f"  Added '{code}' at row {next_empty_row}: {count}")
                
                next_empty_row += 1
        
        # Save the workbook
        wb.save(excel_path)
        
        print(f"\nSpreadsheet updated successfully:")
        print(f"  - Updated existing codes: {updated_count}")
        print(f"  - Added new codes: {added_count}")
        print(f"  - Total codes in spreadsheet: {len(existing_codes) + added_count}")
        print("=" * 80)
        
        return updated_count + added_count
        
    except Exception as e:
        print(f"Error updating Excel file: {e}")
        return 0

def print_code_results(code_counter):
    """Print the code counting results in a formatted way."""
    if not code_counter:
        print("\nNo codes found in any PDF.")
        return
    
    print("\n=== CODE COUNT RESULTS ===")
    print(f"\nTotal unique codes: {len(code_counter)}")
    print(f"Total occurrences: {sum(code_counter.values())}")
    print("\nCode counts (sorted by frequency):")
    print("-" * 40)
    
    for code, count in code_counter.most_common():
        print(f"  {code}: {count}")
    
    print("\nAlphabetical listing:")
    print("-" * 40)
    for code in sorted(code_counter.keys()):
        print(f"  {code}: {code_counter[code]}")

def main():
    if len(sys.argv) < 2 or len(sys.argv) > 4:
        print("Usage: python batch_process.py <folder_path> [excel_file] [--verbose]")
        print("\nThis script will:")
        print("  1. Count three-letter codes from all PDFs in the folder")
        print("  2. Update counts in Excel spreadsheet (if provided)")
        print("  3. Combine PDFs (excluding 'multi-page' files) sorted by first line")
        print("\nArguments:")
        print("  folder_path  - Path to folder containing PDF files")
        print("  excel_file   - (Optional) Path to Excel file to update with counts")
        print("                 If not provided, looks for .xlsx file in folder")
        print("\nOptions:")
        print("  --verbose    - Show detailed processing information")
        print("\nExamples:")
        print("  python batch_process.py ./pdfs")
        print("  python batch_process.py ./pdfs counts.xlsx")
        print("  python batch_process.py ./pdfs counts.xlsx --verbose")
        sys.exit(1)
    
    folder_path = sys.argv[1]
    
    # Parse arguments
    excel_path = None
    verbose = False
    
    for arg in sys.argv[2:]:
        if arg == "--verbose":
            verbose = True
        elif arg.endswith('.xlsx') or arg.endswith('.xlsm'):
            excel_path = arg
    
    # If no excel file specified, look for one in the folder
    if excel_path is None:
        xlsx_files = list(Path(folder_path).glob('*.xlsx'))
        if xlsx_files:
            excel_path = str(xlsx_files[0])
            print(f"Found Excel file: {excel_path}")
    
    # Verify folder exists
    if not os.path.isdir(folder_path):
        print(f"Error: '{folder_path}' is not a valid directory")
        sys.exit(1)
    
    # Count codes from all PDFs
    print("\n" + "=" * 80)
    print("STEP 1: Counting three-letter codes from all PDFs")
    print("=" * 80)
    code_counter = count_codes_in_folder(folder_path, verbose)
    print_code_results(code_counter)
    
    # Update Excel spreadsheet if path is provided
    if excel_path and code_counter:
        print("\n" + "=" * 80)
        print("STEP 2: Updating Excel spreadsheet")
        print("=" * 80)
        update_excel_spreadsheet(excel_path, code_counter, verbose)
        step_num = 3
    else:
        if not excel_path:
            print("\nNo Excel file found or specified - skipping spreadsheet update")
        step_num = 2
    
    # Combine PDFs
    print("\n" + "=" * 80)
    print(f"STEP {step_num}: Combining PDFs alphabetically")
    print("=" * 80)
    output_path = os.path.join(folder_path, "combined_alphabetical.pdf")
    num_pages = combine_pdfs_alphabetically(folder_path, output_path, verbose)
    
    print(f"\n{'=' * 80}")
    print("COMPLETE!")
    print(f"{'=' * 80}")
    print(f"Total codes counted: {sum(code_counter.values())}")
    if excel_path:
        print(f"Excel file updated: {excel_path}")
    print(f"Combined PDF pages: {num_pages}")
    print(f"Combined PDF output: {output_path}")

if __name__ == "__main__":
    main()
